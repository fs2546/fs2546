# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

workbook = load_workbook("courses.xlsx")
students = workbook['students']
times = workbook['time']
def combine():
	combine = workbook.create_sheet(title="combine")
	combine.append(['创建时间','课程名称','学习人数','学习时间'])
	for student in students.values:
		if student[2] != '学习人数':
			for time in times.values:
				if time[1] == student[1]:
					combine.append(list(student) + [time[2]])
	workbook.save('courses.xlsx')



def split():
	combine = workbook['combine']
	split_name = []
	for item in combine.values:
		if item[0] != '创建时间':
			split_name.append(item[0].strftime("%Y"))

	for name in set(split_name):
		temp = Workbook()
		temp.remove(temp.active)
		worksheet = temp.create_sheet(title=name)
		for item_by_year in combine.values: 
			if item_by_year[0] != '创建时间':
				if item_by_year[0].strftime("%Y") == name:
					worksheet.append(item_by_year)
		temp.save(f'{name}.xlsx')
		

if __name__ == '__main__':
	combine()
	split()
